import functools
import glob
import os
import re
import openpyxl as px
from openpyxl.utils.cell import get_column_letter
from .column_config import ColumnConfig, ValueType, Alignment


COLUMN_CONFIG_PATHS = [
    'columns.yml',
    'columns.yaml',
]

IGNORE_FILE_PATH = 'ignore.txt'

THIN_BORDER = px.styles.Border(left=px.styles.Side(border_style='thin'),
                               right=px.styles.Side(border_style='thin'),
                               top=px.styles.Side(border_style='thin'),
                               bottom=px.styles.Side(border_style='thin'))


def variable_config_names(environment: str = None) -> list[str]:
    """Return a list of paths to variable config files.

    :param environment: environment signature such as "dev" or "test"
    :return: list of paths
    """
    if environment is None or environment == '':
        return ['variables.ini']

    return [
        'variables.ini',
        f'variables.{environment}.ini',
    ]


def read_variables(directory_path, environment: str = None) -> dict[str, str]:
    """Read variables from config files.

    :param directory_path: path to the directory which holds config files
    :param environment: environment signature such as "dev" or "test"
    :return: dictionary of variables
    """
    variables = {}
    for file_name in variable_config_names(environment):
        file_path = os.path.join(directory_path, 'config', file_name)
        if os.path.exists(file_path):
            with open(file_path, 'r') as f:
                result = re.findall(r'^(?P<key>[^#].*)=(?P<value>.*)', f.read(), flags=re.MULTILINE)
                for key, value in result:
                    variables[key.strip()] = value.strip()
    return variables


def read_column_config(directory_path) -> ColumnConfig:
    """Read column config from config files.

    :param directory_path: path to the directory which holds config files
    :return: ColumnConfig object
    """
    column_config = ColumnConfig()
    for path in COLUMN_CONFIG_PATHS:
        file_path = os.path.join(directory_path, 'config', path)
        if os.path.exists(file_path):
            column_config.parse(file_path)
            break
    return column_config


def read_ignore_file(directory_path) -> list[str]:
    """Read ignore file.

    :param directory_path: path to the directory which holds config files
    :return: list of file names
    """
    ignore_file_path = os.path.join(directory_path, 'config', IGNORE_FILE_PATH)
    if os.path.exists(ignore_file_path):
        with open(ignore_file_path, 'r') as f:
            return [line.strip() for line in f.readlines()]
    return []


def trim_blank_lines(lines: list[str]) -> list[str]:
    """Remove blank lines from front and back of lines.

    :param lines: list of lines
    :return: list of lines

    >>> trim_blank_lines([' ', '', 'a', '', 'b', 'c', '', ' ', ''])
    ['a', '', 'b', 'c']
    """
    # remove front blank lines
    blank_count = 0
    for i in range(len(lines)):
        if re.match(r"^\s*$", lines[i]):
            blank_count += 1
        else:
            break
    lines = lines[blank_count:]
    # remove back blank lines
    blank_count = 0
    for i in range(1, len(lines) + 1):
        if re.match(r"^\s*$", lines[-i]):
            blank_count += 1
        else:
            break
    if blank_count > 0:
        lines = lines[:-blank_count]
    return lines


class StepItem:
    def __init__(self, title: str = None, step_type: ValueType = ValueType.STRING):
        self.title = title
        self.type = step_type
        self.content_lines = []
        self.content_items = []

    def add_content_line(self, content) -> 'StepItem':
        if self.type == ValueType.STRING:
            if len(self.content_lines) == 0 and re.match(r'^\s*$', content):
                return self
            self.content_lines.append(content)
        elif self.type == ValueType.LIST:
            if len(self.content_items) == 0 and re.match(r'^\s*$', content):
                return self
            self.content_items.append(re.sub(r'^\s*\*\s*', '', content))
        return self

    def get_content(self) -> str | list:
        if self.type == ValueType.STRING:
            self.content_lines = trim_blank_lines(self.content_lines)
            return "\n".join(self.content_lines)
        if self.type == ValueType.LIST:
            self.content_items = trim_blank_lines(self.content_items)
            return self.content_items
        raise ValueError(f'Type {self.type} does not provide content.')


def apply_variables(value, variables: dict) -> str | None:
    """Apply variables to value.

    :param value: value to apply variables
    :param variables: variables
    :return: value with variables applied or None if value is not string

    >>> apply_variables('a{{b}}c', {'b': 'B'})
    'aBc'
    """
    if not isinstance(value, str):
        return value
    for k, v in variables.items():
        value = re.sub(r'{{\s*' + k + r'\s*}}', v, value)
    return value


def build_excel(directory_path, environment: str = None):
    # load column config
    column_config = read_column_config(directory_path)
    list_columns = column_config.list_columns()

    # load variables from ini
    variables = read_variables(directory_path, environment)

    ignore_file_path = os.path.join(directory_path, 'config', IGNORE_FILE_PATH)
    ignore_file_names = []
    if os.path.exists(ignore_file_path):
        with open(ignore_file_path, 'r') as f:
            ignore_file_names =\
                list(filter(lambda x: x != '', map(lambda x: x.strip(), f.readlines())))

    # create new excel book
    wb = px.Workbook()

    # build Excel file
    for scenario_file in glob.glob(directory_path + '/*.md'):
        if os.path.basename(scenario_file) in ignore_file_names:
            continue

        # add Excel sheet
        with open(scenario_file) as f:
            # set name
            name = None
            while True:
                line = f.readline()

                if not line:
                    break

                result = re.match(r'^#[^#]\s*(\S.*)\s*$', line.rstrip())
                if result:
                    name = result.group(1)
                    break

            if not name:
                raise Exception('Title is not set, which must begin with "#" at the top of the file.')

            ws = wb.create_sheet(name)

            # set summary
            has_summary = False
            while True:
                line = f.readline()

                if not line:
                    break

                result = re.match(r'^##\s*Summary\s*$', line)
                if not result:
                    continue
                has_summary = True
                break

            if not has_summary:
                break

            row_index = 1
            cell = ws.cell(row=row_index, column=1)
            cell.value = 'Summary'
            cell.font = px.styles.Font(bold=True)
            row_index += 2

            # read summary lines
            summary_lines = []
            while True:
                line = f.readline()
                if re.match(r'^##\s*(List|Steps|Rows)\s*$', line):
                    break
                summary_lines.append(line.rstrip())
            summary_lines = trim_blank_lines(summary_lines)

            # write summary lines
            for summary_line in summary_lines:
                ws.cell(row=row_index, column=1).value = apply_variables(summary_line, variables)
                row_index += 1

            row_index += 1

            # read steps
            steps = []
            step_dict = {}
            item = None
            while True:
                line = f.readline()

                if not line:
                    if item:
                        step_dict[item.title] = item.get_content()
                    if len(step_dict) > 0:
                        steps.append(step_dict)
                    break

                if re.match(r'^\s*---\s*$', line):
                    if item:
                        step_dict[item.title] = item.get_content()
                        item = None
                    if len(step_dict) > 0:
                        steps.append(step_dict)
                        step_dict = {}
                    continue

                result = re.match(r'^#{3,}\s*(\S*)\s*$', line)
                if result:
                    if item:
                        step_dict[item.title] = item.get_content()
                    title = result.group(1)
                    item = StepItem(
                        title,
                        column_config.type_of(title)
                    )
                    continue

                if item:
                    item.add_content_line(line.rstrip())

        # update steps
        columns = functools.reduce(lambda x, y: x + [z for z in y if z not in x], map(lambda x: x.keys(), steps), [])
        # Copy the previous column value if the step doesn't have the column
        if column_config.duplicate_previous_for_blank:
            for index, step in enumerate(steps):
                step.update({k: v for k, v in steps[index - 1].items() if k not in step})
        for column in list_columns:
            if column in columns:
                index = columns.index(column)
                count = functools.reduce(max, map(lambda x: len(x[column]) if column in x else 0, steps), 0)
                # add numbered column
                for i in range(count - 1, -1, -1):
                    columns.insert(index + 1, f'{column} ({i + 1})')

                # split list column
                for index, step in enumerate(steps):
                    if column in step:
                        for i in range(len(step[column])):
                            step[f'{column} ({i + 1})'] = step[column][i]
                        del step[column]
                # remove original column
                columns.remove(column)

        for column_index, column in enumerate(column_config.prepend_columns.items()):
            columns.insert(column_index, column[0])

        for column in column_config.append_columns:
            columns.append(column)

        # write header
        for column_index, column in enumerate(columns):
            cell = ws.cell(row=row_index, column=column_index + 1)
            cell.value = column
            cell.font = px.styles.Font(bold=True)
            cell.border = THIN_BORDER

            letter = get_column_letter(column_index + 1)

            # arrange column width
            all_conditions = column_config.all_conditions()
            if column in all_conditions:
                condition = all_conditions[column]
                if condition.width:
                    ws.column_dimensions[letter].width = condition.width
                cell.alignment = condition.alignment.excel_alignment()
            else:
                cell.alignment = Alignment.LEFT.excel_alignment()
        row_index += 1

        # write steps
        increment_columns = column_config.increment_columns()

        for index, step in enumerate(steps):
            increment_value = index + 1
            for column in increment_columns:
                step[column] = increment_value

            for column_index, column in enumerate(columns):
                cell = ws.cell(row=row_index, column=column_index + 1)
                if column in step:
                    cell.value = apply_variables(step[column], variables)
                cell.border = THIN_BORDER
                if column in all_conditions:
                    cell.alignment = all_conditions[column].alignment.excel_alignment()
                else:
                    cell.alignment = Alignment.LEFT.excel_alignment()

            row_index += 1

    wb.remove(wb.worksheets[0])

    # save Excel file
    basename = os.path.basename(os.path.abspath(directory_path))
    if environment is None or environment == '':
        filename = basename + '.xlsx'
    else:
        filename = f'{basename}_{environment}.xlsx'
    if not os.path.exists(os.path.join(directory_path, 'output')):
        os.makedirs(os.path.join(directory_path, 'output'))
    wb.save(os.path.join(directory_path, 'output', filename))

    print('Saved', filename)
    return wb


def is_none_or_blank_string(value):
    return value is None or value.strip() == ''
