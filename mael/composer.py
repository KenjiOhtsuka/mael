import os
import re
from abc import ABC, abstractmethod
from enum import Enum

import openpyxl as px
from openpyxl.utils.cell import get_column_letter
from .column_config import ColumnConfig, ValueType, Alignment, Document

import csv
import shutil

def apply_variables(value, variables: dict) -> str | None:
    """Apply variables to value.

    :param value: value to apply variables
    :param variables: variables
    :return: value with variables applied or None if value is not string

    >>> apply_variables('a{{b}}c', {'b': 'B'})
    'aBc'
    """
    if not isinstance(value, str):
        return value
    for k, v in variables.items():
        value = re.sub(r'{{\s*' + k + r'\s*}}', v, value)
    return value


class Composer(ABC):
    def __init__(self):
        pass

    @abstractmethod
    def add_sheet(self, document, column_config, variables, all_conditions, columns, steps):
        pass

    @abstractmethod
    def compose(self, directory_path, environment, basename):
        pass


class OutputFormat(Enum):
    EXCEL = 'excel'
    CSV = 'csv'
    TSV = 'tsv'

    @classmethod
    def build_composer(cls, form) -> Composer:
        lower_name = str(form).lower()
        if cls.EXCEL == form or cls.EXCEL.name.lower() == lower_name:
            return ExcelComposer()
        if cls.CSV == form or cls.CSV.name.lower() == lower_name:
            return CsvComposer()
        if cls.TSV == form or cls.TSV.name.lower() == lower_name:
            return TsvComposer()
        raise ValueError('Unknown format: ' + str(form))


class ExcelComposer(Composer):
    THIN_BORDER = px.styles.Border(left=px.styles.Side(border_style='thin'),
                                   right=px.styles.Side(border_style='thin'),
                                   top=px.styles.Side(border_style='thin'),
                                   bottom=px.styles.Side(border_style='thin'))

    def __init__(self):
        super().__init__()
        self.workbook = px.Workbook()

    def add_sheet(self, document, column_config, variables, all_conditions, columns, steps):
        ws = self.workbook.create_sheet(document.title)
        row_index = 1
        cell = ws.cell(row=row_index, column=1)
        cell.value = 'Summary'
        cell.font = px.styles.Font(bold=True)
        row_index += 2
        # write summary lines
        for summary_line in document.summary_lines:
            ws.cell(row=row_index, column=1).value = apply_variables(summary_line, variables)
            row_index += 1
        row_index += 1

        # write header
        for column_index, column in enumerate(columns):
            cell = ws.cell(row=row_index, column=column_index + 1)
            cell.value = column
            cell.font = px.styles.Font(bold=True)
            cell.border = ExcelComposer.THIN_BORDER

            letter = get_column_letter(column_index + 1)

            # arrange column width
            if column in all_conditions:
                condition = all_conditions[column]
                if condition.width:
                    ws.column_dimensions[letter].width = condition.width
                cell.alignment = condition.alignment.excel_alignment()
            else:
                cell.alignment = Alignment.LEFT.excel_alignment()
        row_index += 1

        # write steps
        increment_columns = column_config.increment_columns()

        for index, step in enumerate(steps):
            increment_value = index + 1
            for column in increment_columns:
                step[column] = increment_value

            for column_index, column in enumerate(columns):
                cell = ws.cell(row=row_index, column=column_index + 1)
                if column in step:
                    cell.value = apply_variables(step[column], variables)
                cell.border = ExcelComposer.THIN_BORDER
                if column in all_conditions:
                    cell.alignment = all_conditions[column].alignment.excel_alignment()
                else:
                    cell.alignment = Alignment.LEFT.excel_alignment()

            row_index += 1

    def compose(self, directory_path, environment, basename):
        self.workbook.remove(self.workbook.worksheets[0])

        # save Excel file
        if environment is None or environment == '':
            filename = basename + '.xlsx'
        else:
            filename = f'{basename}_{environment}.xlsx'
        if not os.path.exists(os.path.join(directory_path, 'output')):
            os.makedirs(os.path.join(directory_path, 'output'))
        self.workbook.save(os.path.join(directory_path, 'output', filename))
        print('Saved', filename)
        return self.workbook


class CsvComposer(Composer):
    def __init__(self, delimiter: str = ','):
        super().__init__()
        self.documents = []
        self.delimiter = delimiter
        self.extension = 'csv'

    def add_sheet(self, document, column_config, variables, all_conditions, columns, steps):
        doc = {
            'title': document.title,
            'summary_lines': document.summary_lines,
            'columns': columns,
            'rows': []
        }

        # write header
        values = [columns]

        # write steps
        increment_columns = column_config.increment_columns()

        for index, step in enumerate(steps):
            increment_value = index + 1
            for column in increment_columns:
                step[column] = increment_value
            values.append(
                [
                    apply_variables(step[column], variables) if column in step else None
                    for column in columns
                ]
            )

        doc['rows'] = values
        self.documents.append(doc)

    def compose(self, directory_path, environment, basename) -> None:
        if environment is None or environment == '':
            dir_name = basename + '_' + self.extension
        else:
            dir_name = f'{basename}_{environment}_' + self.extension

        dir_path = os.path.join(directory_path, 'output', dir_name)
        if os.path.exists(dir_path) and os.path.isdir(dir_path):
            shutil.rmtree(dir_path)
        os.makedirs(dir_path)

        for doc in self.documents:
            file_name = doc['title'] + '.' + self.extension
            file_path = os.path.join(directory_path, 'output', dir_name, file_name)
            with open(file_path, 'w', newline='') as csvfile:
                writer = csv.writer(csvfile, delimiter=self.delimiter)
                writer.writerows(doc['rows'])
                print('Saved', file_name)


class TsvComposer(CsvComposer):
    def __init__(self):
        super().__init__('\t')
        self.extension = 'tsv'
