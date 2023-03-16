from openpyxl.utils.cell import get_column_letter
import functools
import glob
import openpyxl as px
import openpyxl.styles.alignment
import os
import re

THIN_BORDER = px.styles.Border(left=px.styles.Side(border_style='thin'),
                               right=px.styles.Side(border_style='thin'),
                               top=px.styles.Side(border_style='thin'),
                               bottom=px.styles.Side(border_style='thin'))
ALIGNMENT = openpyxl.styles.alignment.Alignment(
    wrap_text=True, vertical='top')


# create enum
class StepType:
    STRING = 1
    LIST = 2


class ColumnCondition:
    def __init__(self, type: StepType = StepType.STRING, width: int = None):
        self.type = type
        self.width = width


column_conditions = {
    'Categories': ColumnCondition(StepType.LIST),
    'Description': ColumnCondition(width=50),
    'Expected': ColumnCondition(width=50),
    'Comment': ColumnCondition(width=50),
}


def main() -> None:
    # get the file directory
    base_dir = os.getcwd()
    test_dir = base_dir + '/test_scenario'

    # read directories in a directory
    for d in os.listdir(test_dir):
        build_scenario(os.path.join(test_dir, d))


def trim_blank_lines(lines: list[str]) -> list[str]:
    # remove front blank lines
    blank_count = 0
    for i in range(len(lines)):
        if re.match(r"^\s*$", lines[i]):
            blank_count += 1
        else:
            break
    lines = lines[blank_count:]
    # remove back blank lines
    blank_count = 0
    for i in range(1, len(lines) + 1):
        if re.match(r"^\s*$", lines[-i]):
            blank_count += 1
        else:
            break
    if blank_count > 0:
        lines = lines[:-blank_count]
    return lines


class StepItem:
    def __init__(self, title: str = None, step_type: StepType = StepType.STRING):
        self.title = title
        self.type = step_type
        self.content_lines = []
        self.content_items = []

    def add_content_line(self, content) -> 'StepItem':
        if self.type == StepType.STRING:
            if len(self.content_lines) == 0 and re.match(r'^\s*$', content):
                return self
            self.content_lines.append(content)
        elif self.type == StepType.LIST:
            if len(self.content_items) == 0 and re.match(r'^\s*$', content):
                return self
            self.content_items.append(re.sub(r'^\s*\*\s*', '', content))
        return self

    def get_content(self) -> str | list:
        if self.type == StepType.STRING:
            self.content_lines = trim_blank_lines(self.content_lines)
            return "\n".join(self.content_lines)
        elif self.type == StepType.LIST:
            self.content_items = trim_blank_lines(self.content_items)
            return self.content_items


def build_scenario(directory_path):
    # load variables from ini
    ini_path = os.path.join(directory_path, 'config.ini')
    variables = {}
    with open(ini_path, 'r') as f:
        result = re.findall(r'(?P<key>.*)=(?P<value>.*)', f.read(), flags=re.MULTILINE)
        for key, value in result:
            variables[key] = value

    list_columns = [k for k, v in column_conditions.items() if v.type == StepType.LIST]

    # create new excel book
    wb = px.Workbook()

    # build Excel file
    for scenario_file in glob.glob(directory_path + '/*.md'):
        print(scenario_file)

        # add excel sheet
        with open(scenario_file) as f:
            # set name
            name = None
            while True:
                line = f.readline()

                if not line:
                    break

                result = re.match(r'^#\s*(\S.*)\s*$', line)
                if result:
                    name = result.group(1)
                    break

            if not name:
                break

            ws = wb.create_sheet(name)

            # set summary
            has_summary = False
            while True:
                line = f.readline()

                if not line:
                    break

                result = re.match(r'^##\s*Summary\s*$', line)
                if not result:
                    continue
                has_summary = True
                break

            if not has_summary:
                break

            row_index = 1
            cell = ws.cell(row=row_index, column=1)
            cell.value = 'Summary'
            cell.font = px.styles.Font(bold=True)
            row_index += 2

            # read summary lines
            summary_lines = []
            while True:
                line = f.readline()
                if re.match(r'^##\s*Steps\s*$', line):
                    break
                else:
                    summary_lines.append(line.rstrip())
            summary_lines = trim_blank_lines(summary_lines)
            print(summary_lines)

            # write summary lines
            for summary_line in summary_lines:
                ws.cell(row=row_index, column=1).value = summary_line
                row_index += 1

            row_index += 1

            # read steps
            steps = []
            step_dict = {}
            item = None
            while True:
                line = f.readline()
                if not line:
                    if item:
                        step_dict[item.title] = item.get_content()
                    if len(step_dict) > 0:
                        steps.append(step_dict)
                    break

                if re.match(r'^\s*---\s*$', line):
                    if item:
                        step_dict[item.title] = item.get_content()
                        item = None
                    if len(step_dict) > 0:
                        steps.append(step_dict)
                        step_dict = {}
                    continue

                result = re.match(r'^###\s*(.*)\s*$', line)
                if result:
                    if item:
                        step_dict[item.title] = item.get_content()
                    title = result.group(1)
                    item = StepItem(
                        title,
                        column_conditions[title].type if title in column_conditions else StepType.STRING
                    )
                    continue

                if item:
                    item.add_content_line(line.rstrip())

        # write steps
        columns = functools.reduce(lambda x, y: x + [z for z in y if z not in x], map(lambda x: x.keys(), steps), [])
        for column in list_columns:
            if column in columns:
                index = columns.index(column)
                count = functools.reduce(max, map(lambda x: len(x[column]) if column in x else 0, steps), 0)
                # add numbered column
                for i in range(count - 1, -1, -1):
                    columns.insert(index + 1, f'{column} ({i + 1})')
                # split list column
                for step in steps:
                    if column in step:
                        for i in range(len(step[column])):
                            step[f'{column} ({i + 1})'] = step[column][i]
                        del step[column]
                # remove original column
                columns.remove(column)

        columns.insert(0, "No.")
        columns.append("Timestamp")
        columns.append("Result")
        columns.append("Comment")

        # write header
        for column_index, column in enumerate(columns):
            cell = ws.cell(row=row_index, column=column_index + 1)
            cell.value = column
            cell.font = px.styles.Font(bold=True)
            cell.border = THIN_BORDER

            letter = get_column_letter(column_index + 1)

            # arrange column width
            if column in column_conditions:
                condition = column_conditions[column]
                if condition.width:
                    ws.column_dimensions[letter].width = condition.width
        row_index += 1

        # write steps
        for index, step in enumerate(steps):
            step['No.'] = index + 1

            for column_index, column in enumerate(columns):
                cell = ws.cell(row=row_index, column=column_index + 1)
                if column in step:
                    cell.value = step[column]
                cell.border = THIN_BORDER
                cell.alignment = ALIGNMENT

            row_index += 1

    # save Excel file
    wb.save(os.path.join(directory_path, f'{os.path.basename(directory_path)}.xlsx'))
    return wb


if __name__ == '__main__':
    # TODO:
    #   * replace variable
    #   * output variables
    #   * change font for code part which is surrounded by ```
    main()
